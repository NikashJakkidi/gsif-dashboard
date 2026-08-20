import json
import subprocess
from datetime import datetime
from pathlib import Path

import openpyxl

GSIF_TRACKER_PATH = r"C:\Users\nikas\OneDrive - University of Florida\GSIF General Body Archive (2026)\Portfolio Tools\GSIF Official Portfolio Tracker.xlsx"
SP1000_TRACKER_PATH = r"C:\Users\nikas\OneDrive - University of Florida\GSIF General Body Archive (2026)\Portfolio Tools\S&P 1000 Tracker.xlsx"
OUTPUT_HTML = Path(__file__).parent / "index.html"

TICKER, COMPANY, SECTOR, SUBIND, VERTICAL = 1, 3, 4, 6, 7
END_PRICE, END_VALUE, END_VERT_W, END_SECTOR_W, END_FUND_W = 18, 19, 20, 21, 22
LABEL_COL, VAL_A_COL, VAL_B_COL = 24, 25, 26

VERT_SECTOR_COL, VERT_VERTICAL_COL, VERT_WEIGHT_COL = 11, 12, 13
BENCH_SECTOR_COL, BENCH_WEIGHT_COL = 15, 16


def load_gsif_data(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Current (Hardcode)"]

    holdings = []
    summary = {}
    sector_weights = {}
    in_sector_table = False

    for row in range(3, ws.max_row + 1):
        ticker = ws.cell(row, TICKER).value
        if ticker and ticker != "Cash":
            holdings.append({
                "ticker": ticker,
                "company": ws.cell(row, COMPANY).value,
                "sector": ws.cell(row, SECTOR).value,
                "subindustry": ws.cell(row, SUBIND).value,
                "vertical": ws.cell(row, VERTICAL).value,
                "end_price": ws.cell(row, END_PRICE).value,
                "end_value": ws.cell(row, END_VALUE).value or 0,
                "end_vertical_weight": ws.cell(row, END_VERT_W).value,
                "end_sector_weight": ws.cell(row, END_SECTOR_W).value,
                "end_fund_weight": ws.cell(row, END_FUND_W).value,
            })

        label = ws.cell(row, LABEL_COL).value
        val_a = ws.cell(row, VAL_A_COL).value
        val_b = ws.cell(row, VAL_B_COL).value

        if label is None:
            in_sector_table = False
            continue
        if label == "Sector":
            in_sector_table = True
            continue
        if in_sector_table:
            sector_weights[label] = (val_a, val_b)
        else:
            summary[label] = (val_a, val_b)

    return holdings, summary, sector_weights


def load_sp1000_data(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Hardcode Current S&P 1000"]

    vertical_weight_in_sector = {}
    row = 3
    while True:
        sector = ws.cell(row, VERT_SECTOR_COL).value
        if sector is None:
            break
        vertical = ws.cell(row, VERT_VERTICAL_COL).value
        weight = ws.cell(row, VERT_WEIGHT_COL).value
        vertical_weight_in_sector[(sector, vertical)] = weight
        row += 1

    sector_weight_in_benchmark = {}
    row = 3
    while True:
        sector = ws.cell(row, BENCH_SECTOR_COL).value
        if sector is None:
            break
        weight = ws.cell(row, BENCH_WEIGHT_COL).value
        sector_weight_in_benchmark[sector] = weight
        row += 1

    return vertical_weight_in_sector, sector_weight_in_benchmark


def to_date_str(value):
    return value.strftime("%Y-%m-%d") if hasattr(value, "strftime") else str(value)


def build_dashboard_data():
    holdings, summary, sector_weights = load_gsif_data(GSIF_TRACKER_PATH)
    vert_weight_bench, sector_weight_bench = load_sp1000_data(SP1000_TRACKER_PATH)

    trade_date, as_of_date = summary["As of Date"]
    equity_start, equity_end = summary["Fund Equity Value"]
    _, cash_end = summary["Estimated Cash"]
    aum_start, aum_end = summary["Estimated Total AUM"]

    sectors = []
    for sector_name, (w_start, w_end) in sector_weights.items():
        bench_w = sector_weight_bench.get(sector_name) or 0
        sec_holdings = [h for h in holdings if h["sector"] == sector_name]
        sector_aum = sum(h["end_value"] for h in sec_holdings)

        verticals_in_sector = {v for (s, v) in vert_weight_bench if s == sector_name}
        verticals_in_sector |= {h["vertical"] for h in sec_holdings}

        vertical_rows = []
        for v in verticals_in_sector:
            v_value = sum(h["end_value"] for h in sec_holdings if h["vertical"] == v)
            port_w = (v_value / sector_aum) if sector_aum else 0
            bench_vw = vert_weight_bench.get((sector_name, v)) or 0
            vertical_rows.append({
                "vertical": v,
                "portfolio_weight_in_sector": port_w,
                "benchmark_weight_in_sector": bench_vw,
                "diff": port_w - bench_vw,
            })
        vertical_rows.sort(key=lambda r: r["portfolio_weight_in_sector"], reverse=True)

        holdings_table = sorted(
            (
                {
                    "ticker": h["ticker"],
                    "company": h["company"],
                    "vertical": h["vertical"],
                    "weight_in_sector": h["end_sector_weight"] or 0,
                    "weight_in_vertical": h["end_vertical_weight"] or 0,
                }
                for h in sec_holdings
            ),
            key=lambda r: (r["vertical"] or "", -r["weight_in_vertical"]),
        )

        sectors.append({
            "name": sector_name,
            "weight_portfolio": w_end or 0,
            "weight_benchmark": bench_w,
            "active_weight": (w_end or 0) - bench_w,
            "aum": sector_aum,
            "num_holdings": len(sec_holdings),
            "verticals": vertical_rows,
            "holdings": holdings_table,
        })

    sectors.sort(key=lambda s: s["weight_portfolio"], reverse=True)

    return {
        "as_of_date": to_date_str(as_of_date),
        "trade_date": to_date_str(trade_date),
        "fund": {
            "total_aum": aum_end,
            "total_aum_start": aum_start,
            "equity_value": equity_end,
            "equity_value_start": equity_start,
            "cash": cash_end,
            "num_holdings": len(holdings),
        },
        "sectors": sectors,
    }


def render_html(data):
    template_path = Path(__file__).parent / "dashboard_template.html"
    template = template_path.read_text(encoding="utf-8")
    return template.replace("__DASHBOARD_DATA__", json.dumps(data))


def git_publish(project_dir):
    def run(cmd):
        return subprocess.run(cmd, cwd=project_dir, capture_output=True, text=True)

    if not (project_dir / ".git").exists():
        run(["git", "init"])
        run(["git", "branch", "-M", "main"])

    run(["git", "add", "-A"])
    diff = run(["git", "diff", "--cached", "--quiet"])
    if diff.returncode == 0:
        print("No changes since last run — nothing to publish.")
        return

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    commit = run(["git", "commit", "-m", f"Update dashboard — {stamp}"])
    if commit.returncode != 0:
        print("git commit failed:")
        print(commit.stdout, commit.stderr)
        return

    push = run(["git", "push", "-u", "origin", "main"])
    if push.returncode != 0:
        print("git push failed:")
        print(push.stdout, push.stderr)
        print("If this is the first run, make sure the GitHub remote 'origin' is configured.")
        return

    print("Published.")


def main():
    print("Reading workbooks...")
    data = build_dashboard_data()
    print(f"  Fund AUM: ${data['fund']['total_aum']:,.2f}  |  Holdings: {data['fund']['num_holdings']}  |  As of {data['as_of_date']}")

    html = render_html(data)
    OUTPUT_HTML.write_text(html, encoding="utf-8")
    print(f"Wrote {OUTPUT_HTML}")

    git_publish(OUTPUT_HTML.parent)


if __name__ == "__main__":
    main()
